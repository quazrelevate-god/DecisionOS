"""Epic 10 Sprint 12 -- the exit gate (T10-12.9) + go-live QA sign-off (T10-12.10).

These read the Epic 10 tracker (the authoritative record of which testing
dimension is green) and:

  .9  map every Testing-Plan phase A-J (docs Epic 1 "Testing Plan") to a concrete
      pass/fail gate, and assert the AUTOMATABLE phases have met their exit
      criteria (their Epic 10 sprints are Done). The non-automatable phases
      (load, chaos, backup/DR, alpha/beta/GA) are explicitly attributed to ops
      (Epic 1 Sprint 5 / rollout), not silently claimed.

  .10 the formal QA sign-off: every AUTOMATED testing dimension (Epic 10 S1-S11
      + the S12 E2E journeys) is green. The Go-Live Checklist "Testing (A-J)"
      row stays pending until the manual/ops phases land (Epic 1 S5) -- this test
      records the automated half of that sign-off honestly.
"""
from collections import defaultdict
from pathlib import Path

import openpyxl

_TRACKER = Path(__file__).resolve().parent.parent.parent / "docs" / "DecisionOS_Epic10_Testing.xlsx"


def _load():
    """Return (per_sprint {sprint:int -> (done,total)}, per_task {id -> status})."""
    wb = openpyxl.load_workbook(_TRACKER, read_only=True)
    ws = wb["Backlog"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {v: i for i, v in enumerate(hdr)}
    done = defaultdict(int); total = defaultdict(int); task = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[idx["ID"]] is None:
            continue
        sp = r[idx["Sprint"]]
        st = str(r[idx["Status"]]).strip().lower()
        total[sp] += 1
        if st == "done":
            done[sp] += 1
        task[str(r[idx["ID"]])] = st
    wb.close()
    return {s: (done[s], total[s]) for s in total}, task


# Testing-Plan phase -> {epic10 sprint numbers that satisfy it, automatable?, owner}
# Automatable phases are gated by this epic; the rest are ops/rollout (Epic 1 S5).
_PHASES = {
    "A": {"name": "Automated (unit/integration/regression)", "sprints": [1, 2, 11],
          "e2e": True, "automatable": True, "owner": "Epic 10"},
    "B": {"name": "Multi-tenant isolation", "sprints": [8],
          "automatable": True, "owner": "Epic 10"},
    "C": {"name": "Security / RBAC audit", "sprints": [3],
          "automatable": True, "owner": "Epic 10 (RBAC); pentest = manual Epic 1 S0"},
    "D": {"name": "Load / performance", "sprints": [],
          "automatable": False, "owner": "Epic 1 S5 (k6/Locust on staging)"},
    "E": {"name": "Resilience / concurrency / chaos", "sprints": [10],
          "automatable": True, "owner": "Epic 10 (concurrency); chaos = manual Epic 1 S5"},
    "F": {"name": "Migration & backup/DR & tenant-deletion", "sprints": [8],
          "automatable": True, "owner": "Epic 10 (tenant-deletion); migration/backup = manual Epic 1 S5"},
    "G": {"name": "Internal alpha", "sprints": [], "automatable": False, "owner": "manual (internal team)"},
    "H": {"name": "Closed beta", "sprints": [], "automatable": False, "owner": "manual (design partners)"},
    "I": {"name": "Staged open beta", "sprints": [], "automatable": False, "owner": "manual (tiered rollout)"},
    "J": {"name": "General availability", "sprints": [], "automatable": False, "owner": "manual (go-live)"},
}

_S12_JOURNEYS = [f"T10-12.{n}" for n in range(1, 9)]   # .1-.8 (the E2E journeys; .9/.10 are the gates)


# ===========================================================================
# T10-12.9  Phase A-J exit-criteria gate
# ===========================================================================
def test_phase_map_covers_A_through_J():
    assert set(_PHASES) == set("ABCDEFGHIJ"), "the phase gate must map every Testing-Plan phase A-J"


def test_automatable_phases_met_their_exit_criteria():
    per_sprint, per_task = _load()
    failures = []
    for ph, spec in _PHASES.items():
        if not spec["automatable"]:
            continue
        for s in spec["sprints"]:
            d, t = per_sprint.get(s, (0, 0))
            if t == 0 or d != t:
                failures.append(f"phase {ph} ({spec['name']}): sprint S{s} is {d}/{t}, not complete")
        if spec.get("e2e"):
            open_j = [j for j in _S12_JOURNEYS if per_task.get(j) != "done"]
            if open_j:
                failures.append(f"phase {ph}: E2E journeys not all done: {open_j}")
    assert not failures, "automatable Testing-Plan phases not at exit criteria:\n  " + "\n  ".join(failures)


def test_non_automatable_phases_are_attributed_not_silently_claimed():
    # every manual/ops phase must name its owner so the gate never reads as "all green"
    for ph, spec in _PHASES.items():
        if not spec["automatable"]:
            assert spec["owner"] and "manual" in spec["owner"].lower() or "Epic 1" in spec["owner"], \
                f"phase {ph} must be attributed to a manual/ops owner"


# ===========================================================================
# T10-12.10  Go-live testing sign-off (automated dimensions)
# ===========================================================================
def test_all_automated_testing_dimensions_green():
    """The QA sign-off gate: Epic 10 S1-S11 fully Done + the S12 E2E journeys Done.
    This is the automated half of the Go-Live 'Testing' checklist row; the manual
    phases (load/backup/beta) remain with Epic 1 S5."""
    per_sprint, per_task = _load()
    not_green = []
    for s in range(1, 12):   # S1..S11
        d, t = per_sprint.get(s, (0, 0))
        if t == 0 or d != t:
            not_green.append(f"S{s}={d}/{t}")
    open_j = [j for j in _S12_JOURNEYS if per_task.get(j) != "done"]
    assert not not_green, f"automated testing dimensions not all green: {not_green}"
    assert not open_j, f"S12 E2E journeys not all done: {open_j}"
